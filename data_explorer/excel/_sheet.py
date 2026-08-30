"""Small helpers for building consistent, filterable Excel sheets."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13, color="1F3864")
_NOTE_FONT = Font(italic=True, color="595959")

_MAX_W = 70
_MIN_W = 9


def add_table(wb, name: str, headers: list[str], rows: list[list], *,
              title: str = "", note: str = "", widths: dict[str, int] | None = None,
              freeze: str = "A2") -> None:
    ws = wb.create_sheet(name[:31])
    r0 = 1
    if title:
        ws.cell(row=r0, column=1, value=title).font = _TITLE_FONT
        r0 += 1
    if note:
        c = ws.cell(row=r0, column=1, value=note)
        c.font = _NOTE_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=max(3, len(headers)))
        r0 += 1
    hdr_row = r0
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=j, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
    for i, row in enumerate(rows, start=hdr_row + 1):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    last_col = get_column_letter(len(headers))
    last_row = hdr_row + len(rows)
    if rows:
        tbl = Table(displayName=_safe_name(name), ref=f"A{hdr_row}:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)
    else:
        ws.auto_filter.ref = f"A{hdr_row}:{last_col}{hdr_row}"

    # column widths
    for j, h in enumerate(headers, start=1):
        col = get_column_letter(j)
        w = (widths or {}).get(h)
        if w is None:
            sample = [str(h)] + [str(row[j - 1]) for row in rows[:200] if row[j - 1] is not None]
            w = min(_MAX_W, max(_MIN_W, max((len(s) for s in sample), default=_MIN_W) + 2))
        ws.column_dimensions[col].width = w

    ws.freeze_panes = ws[f"A{hdr_row + 1}"]
    return ws


def add_dropdown(ws, col_letter: str, options: list[str], *, first_row: int = 2, last_row: int = 2000) -> None:
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def _safe_name(s: str) -> str:
    return "".join(ch if ch.isalnum() else "_" for ch in s)[:31]
