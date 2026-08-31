"""Build data/data_explorer/exports/MNP_ESCO_ONET_DATA_EXPLORER.xlsx from
the reference DB + the MNP snapshot + the analysis outputs.
"""

from __future__ import annotations

import sqlite3

from openpyxl import Workbook

from data_explorer import config
from data_explorer.analysis import dimensions, quality
from data_explorer.excel._sheet import add_dropdown, add_table
from data_explorer.human_lab import schema as hl_schema
from data_explorer.io import log
from data_explorer.mnp_snapshot import load_mnp_careers

OUT = config.EXPORT_DIR / "MNP_ESCO_ONET_DATA_EXPLORER.xlsx"


def _q(conn, sql, *a):
    cur = conn.execute(sql, a)
    return [d[0] for d in cur.description], cur.fetchall()


def _focus_socs(conn) -> list[str]:
    """O*NET-SOC codes that matter for the current MNP KB: every code a
    mapping_review candidate points at. Keeps the workbook a working tool,
    not a 400k-row dump — the full data stays in reference.sqlite."""
    return sorted({r[0] for r in conn.execute(
        "SELECT DISTINCT external_id FROM mapping_review WHERE source_system='onet' AND external_id <> ''")})


def _focus_esco_uris(conn) -> list[str]:
    return sorted({r[0] for r in conn.execute(
        "SELECT DISTINCT external_id FROM mapping_review WHERE source_system='esco' AND external_id <> ''")})


def _inlist(items: list[str]) -> str:
    return "(" + ",".join("?" * len(items)) + ")"


def build(dest=None) -> None:
    from pathlib import Path
    if not config.REFERENCE_DB.exists():
        raise SystemExit("reference.sqlite not built — run: python -m data_explorer.cli build")
    dest = Path(dest) if dest else OUT
    conn = sqlite3.connect(config.REFERENCE_DB)
    wb = Workbook()
    wb.remove(wb.active)

    _readme(wb)
    _sources(wb, conn)
    _mnp_careers(wb)
    _esco_occupations(wb, conn)
    _onet_occupations(wb, conn)
    _crosswalk(wb, conn)
    _onet_ratings(wb, conn)
    _onet_work_styles_full(wb, conn)
    _esco_occ_skills(wb, conn)
    _career_explorer(wb, conn)
    _dimension_analysis(wb, conn)
    _data_quality(wb, conn)
    _mapping_review(wb, conn)
    _human_lab_templates(wb)
    _golden_export(wb)

    conn.close()
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    log(f"  wrote {dest}  ({len(wb.sheetnames)} sheets)")


# --------------------------------------------------------------------------
def _readme(wb) -> None:
    ws = wb.create_sheet("00_README")
    lines = [
        ("MNP ESCO / O*NET DATA EXPLORER", True),
        ("A read-only research view of the real ESCO + O*NET data for «МОЖУ: Мій Напрям». Generated — do not hand-edit the data sheets.", False),
        ("", False),
        (f"O*NET {config.ONET_RELEASE_LABEL}  ·  O*NET Work Values {config.ONET_WORK_VALUES_LABEL}  ·  ESCO {config.ESCO_LABEL} (en+uk)  ·  official ESCO<->O*NET crosswalk", False),
        ("Rebuild: pip install -r requirements-datalab.txt ; python -m data_explorer.cli download ; python -m data_explorer.cli build ; python -m data_explorer.cli excel", False),
        ("", False),
        ("Sheets", True),
        ("01_SOURCES / 02_SOURCE_VERSIONS   — every dataset, version, sha256, licence, attribution", False),
        ("10_MNP_CAREERS                    — the MNP Career KB (5 ACTIVE alpha careers) + external mappings", False),
        ("11_ESCO_OCCUPATIONS / 12_ONET_OCCUPATIONS / 13_CAREER_CROSSWALK", False),
        ("20_ONET_RATINGS                   — ALL O*NET dimensions (filter by table_key / element)", False),
        ("22_ONET_WORK_STYLES_FULL          — all 21 work styles + discriminative power (brief §14)", False),
        ("23_ESCO_OCC_SKILLS                — ESCO occupation<->skill (essential/optional), en+uk", False),
        ("40_CAREER_EXPLORER                — filter by mnp_career: SOURCE FACT vs MNP INTERPRETATION", False),
        ("41_DIMENSION_ANALYSIS / 42_DATA_QUALITY", False),
        ("50_MAPPING_REVIEW                 — MNP<->external candidates; set review_state / mapping_type (dropdowns)", False),
        ("55_PERSON_PROFILE_TEMPLATE / 58_EXPECTED_RESULT_TEMPLATE  — hand-entry with dropdowns", False),
        ("60_GOLDEN_EXPORT                  — exported Human Expected Results", False),
        ("", False),
        ("RULES: UNKNOWN != 0 (a missing value is a blank cell, never 0). No AI anywhere. No Ukrainian market data. No Lightcast.", False),
    ]
    from openpyxl.styles import Font
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=12, color="1F3864")
    ws.column_dimensions["A"].width = 120


def _sources(wb, conn) -> None:
    h, rows = _q(conn, "SELECT source_label, kind, version, official_url, file, sha256, license FROM stg_source ORDER BY kind")
    add_table(wb, "01_SOURCES", h, [list(r) for r in rows],
              title="Source datasets", note="Downloaded (never scraped); sha256 recorded; data/ is gitignored.")
    h, rows = _q(conn, "SELECT source_label, attribution FROM stg_source ORDER BY source_label")
    add_table(wb, "02_SOURCE_VERSIONS", ["source_label", "attribution (show wherever this data reaches an end user)"],
              [list(r) for r in rows], widths={"attribution (show wherever this data reaches an end user)": 110})


def _mnp_careers(wb) -> None:
    careers = load_mnp_careers()
    rows = []
    for c in careers:
        rows.append([
            c.code, c.canonical_name_uk, c.canonical_name_en, c.status, c.career_profile_version,
            "yes" if c.market_data_limited else "no",
            "; ".join(f"{s['skill_en']} ({s['requirement_type']}/{s['importance']})" for s in c.skill_requirements),
            "; ".join(f"{r['category']}: {r['description']} [{r['hardness']}]" for r in c.requirements),
            "; ".join(f"{a['group']}.{a['key']}={a['value_numeric']}" for a in c.attributes),
            "; ".join(f"{m['source_system']}:{m['external_id']} ({m['mapping_type']})" for m in c.external_mappings) or "(UNMAPPED)",
        ])
    add_table(wb, "10_MNP_CAREERS",
              ["code", "name_uk", "name_en", "status", "profile_version", "market_data_limited",
               "skill_requirements", "requirements", "attributes (MNP curated)", "external_mappings"],
              rows, title="MNP Career Knowledge Base (alpha seed)",
              note="Founder Decision #16: `code` is the only identity. All content is `mnp_editorial_v1`. "
                   "external_mappings currently empty — candidates are on 50_MAPPING_REVIEW.",
              widths={"skill_requirements": 60, "requirements": 60, "attributes (MNP curated)": 40, "external_mappings": 40})


def _esco_occupations(wb, conn) -> None:
    h, rows = _q(conn,
        "SELECT code, isco_group, preferred_label_en, preferred_label_uk, "
        "(SELECT count(*) FROM esco_occupation_skill s WHERE s.occupation_uri=o.uri AND s.relation_type='essential') AS essential_skills, "
        "(SELECT count(*) FROM esco_occupation_skill s WHERE s.occupation_uri=o.uri AND s.relation_type='optional') AS optional_skills, "
        "status, regulated_note_en, uri "
        "FROM esco_occupation o ORDER BY isco_group, preferred_label_en")
    add_table(wb, "11_ESCO_OCCUPATIONS", h, [list(r) for r in rows],
              title=f"ESCO {config.ESCO_LABEL} occupations (en + uk)",
              widths={"preferred_label_en": 45, "preferred_label_uk": 45, "regulated_note_en": 40, "uri": 55})


def _onet_occupations(wb, conn) -> None:
    h, rows = _q(conn,
        "SELECT soc, title, job_zone, "
        "(SELECT count(*) FROM onet_rating r WHERE r.soc=o.soc) AS rating_rows, "
        "(SELECT count(*) FROM onet_task t WHERE t.soc=o.soc) AS tasks, "
        "(SELECT count(*) FROM onet_work_value w WHERE w.soc=o.soc) AS work_values, "
        "description FROM onet_occupation o ORDER BY soc")
    add_table(wb, "12_ONET_OCCUPATIONS", h, [list(r) for r in rows],
              title=f"O*NET {config.ONET_RELEASE_LABEL} occupations",
              widths={"title": 45, "description": 90})


def _crosswalk(wb, conn) -> None:
    h, rows = _q(conn,
        "SELECT esco_isco_code, esco_isco_title, code_level, onet_soc, onet_title, "
        "mapping_relation, esco_occupation_uri FROM xwalk_esco_onet ORDER BY esco_isco_code, onet_soc")
    add_table(wb, "13_CAREER_CROSSWALK", h, [list(r) for r in rows],
              title="Official ESCO<->O*NET-SOC crosswalk (O*NET Resource Center)",
              note="Flat many-to-many. mapping_relation is 'unspecified' — the O*NET-hosted file carries no "
                   "exact/close/broad/narrow semantics, and it is NEVER promoted to 'exact'.",
              widths={"esco_isco_title": 40, "onet_title": 45, "esco_occupation_uri": 55})


def _onet_ratings(wb, conn) -> None:
    socs = _focus_socs(conn)
    h, rows = _q(conn,
        "SELECT r.soc, o.title, r.table_key, r.family, r.element_id, r.element_name, r.scale_id, "
        "r.category, r.raw_value, r.normalized_value "
        "FROM onet_rating r JOIN onet_occupation o ON o.soc=r.soc "
        f"WHERE r.soc IN {_inlist(socs)} ORDER BY r.soc, r.table_key, r.element_id", *socs)
    add_table(wb, "20_ONET_RATINGS", h, [list(r) for r in rows],
              title="ALL O*NET dimensions for the focus occupations (RAW + normalized distinct)",
              note="Focus set = every O*NET-SOC a 50_MAPPING_REVIEW candidate points at "
                   f"({len(socs)} occupations). Filter by table_key (interest / work_style / ability / "
                   "skill_essential / skill_transferable / knowledge / work_activity / work_context / education / "
                   "training_experience). normalized_value is blank for category scales. A missing value is an "
                   "ABSENT ROW, never 0. Full ~400k-row set: reference.sqlite.",
              widths={"element_name": 42, "family": 40, "title": 34})


def _onet_work_styles_full(wb, conn) -> None:
    stats = dimensions.family_stats(conn, "work_style", scale_id="WI")
    rows = [[s.element_id, s.element_name, s.n_occupations, s.coverage, s.mean, s.stdev, s.variance,
             s.minimum, s.maximum, "YES" if s.selected_by_mnp else ""] for s in
            sorted(stats, key=lambda s: s.stdev or 0, reverse=True)]
    add_table(wb, "22_ONET_WORK_STYLES_FULL",
              ["element_id", "element_name", "n_occupations", "coverage", "mean_norm",
               "stdev_norm (discriminative power)", "variance", "min", "max", "MNP_selected"],
              rows, title="All 21 O*NET Work Styles vs the MNP-selected subset (brief §14)",
              note="stdev across occupations = how much the dimension separates jobs. See "
                   "findings/WORK_STYLE_DIFFERENTIATION_FINDING_V1.md. Production code unchanged.")


def _esco_occ_skills(wb, conn) -> None:
    uris = _focus_esco_uris(conn)
    h, rows = _q(conn,
        "SELECT o.preferred_label_en AS occupation_en, o.preferred_label_uk AS occupation_uk, "
        "os.relation_type, s.preferred_label_en AS skill_en, s.preferred_label_uk AS skill_uk, "
        "s.skill_type, s.reuse_level "
        "FROM esco_occupation_skill os JOIN esco_occupation o ON o.uri=os.occupation_uri "
        f"JOIN esco_skill s ON s.uri=os.skill_uri WHERE os.occupation_uri IN {_inlist(uris)} "
        "ORDER BY o.preferred_label_en, os.relation_type", *uris)
    add_table(wb, "23_ESCO_OCC_SKILLS", h, [list(r) for r in rows],
              title="ESCO occupation <-> skill (essential / optional), en + uk — focus occupations",
              note=f"Focus set = every ESCO occupation a 50_MAPPING_REVIEW candidate points at ({len(uris)}). "
                   "Candidate source for MNP skill requirements (Explorer only — Founder Decision #17). "
                   "Full 126k-row set: reference.sqlite.",
              widths={"occupation_en": 38, "occupation_uk": 38, "skill_en": 38, "skill_uk": 38})


def _career_explorer(wb, conn) -> None:
    """Long, filter-by-career table: for each MNP career, its crosswalk +
    the O*NET/ESCO facts and whether MNP uses them (brief §22)."""
    conn.row_factory = sqlite3.Row
    from data_explorer.explorer import views
    rows: list[list] = []
    for c in load_mnp_careers():
        rows.append([c.code, "identity", "mnp", "name", c.canonical_name_en, c.canonical_name_uk, ""])
        # candidate crosswalk targets from mapping_review
        for mr in conn.execute("SELECT source_system, external_id, external_label, signal, score, proposed_mapping_type "
                               "FROM mapping_review WHERE mnp_code=? ORDER BY source_system, score DESC", (c.code,)):
            rows.append([c.code, "mapping_candidate", mr["source_system"], mr["signal"],
                         f'{mr["external_id"]}  {mr["external_label"] or ""}', "",
                         f'score={mr["score"]} type={mr["proposed_mapping_type"]} -> review on 50_MAPPING_REVIEW'])
        # if there is a confirmed/candidate O*NET soc, show used-vs-ignored
        soc_row = conn.execute("SELECT external_id FROM mapping_review WHERE mnp_code=? AND source_system='onet' "
                               "AND signal='title_match' ORDER BY score DESC LIMIT 1", (c.code,)).fetchone()
        if soc_row:
            soc = soc_row["external_id"]
            uvi = views.mnp_used_vs_ignored(conn, soc)
            for fam, items in uvi["used"].items():
                for it in items:
                    rows.append([c.code, "onet_dimension_USED", "onet", fam, it["element_name"], "", f"(best-guess SOC {soc})"])
            for fam, items in uvi["ignored"].items():
                for it in items[:8]:
                    rows.append([c.code, "onet_dimension_IGNORED", "onet", fam, it["element_name"], "", f"(best-guess SOC {soc})"])
    conn.row_factory = None
    add_table(wb, "40_CAREER_EXPLORER",
              ["mnp_career", "section", "source", "key", "value_en", "value_uk", "note"],
              rows, title="Career Explorer — filter `mnp_career`, read SOURCE FACT vs MNP INTERPRETATION",
              note="Static generated view (brief §22 — reliability over Excel formulas). "
                   "onet_dimension rows use the top title-match SOC as a best guess until a human confirms a crosswalk.",
              widths={"value_en": 45, "value_uk": 35, "note": 45})


def _dimension_analysis(wb, conn) -> None:
    rows = []
    for tk in ("interest", "work_style", "ability", "skill_essential", "knowledge", "work_activity", "work_context"):
        for s in dimensions.family_stats(conn, tk):
            rows.append([tk, s.element_id, s.element_name, s.scale_id, s.n_occupations, s.coverage,
                         s.mean, s.stdev, s.minimum, s.maximum, "YES" if s.selected_by_mnp else ""])
    add_table(wb, "41_DIMENSION_ANALYSIS",
              ["table_key", "element_id", "element_name", "scale_id", "n_occupations", "coverage",
               "mean_norm", "stdev_norm", "min", "max", "MNP_selected"],
              rows, title="Per-dimension coverage / distribution / discriminative power (research metrics only)",
              note="These NEVER change production methodology (brief §13).",
              widths={"element_name": 40})


def _data_quality(wb, conn) -> None:
    rep = quality.run(conn)
    add_table(wb, "42_DATA_QUALITY", ["severity", "check", "count"],
              [[s, d, c] for s, d, c in rep.checks],
              title="Data quality — any `error` with count>0 blocks Golden use",
              widths={"check": 90})


def _mapping_review(wb, conn) -> None:
    h, rows = _q(conn,
        "SELECT mnp_code, mnp_name_en, source_system, external_id, external_label, signal, score, "
        "proposed_mapping_type, review_state, confidence, reviewer, note, source_version FROM mapping_review "
        "ORDER BY mnp_code, source_system, score DESC")
    ws = add_table(wb, "50_MAPPING_REVIEW", h, [list(r) for r in rows],
                   title="MNP <-> external mapping CANDIDATES — set review_state + mapping_type by hand",
                   note="Candidates only. Never auto-written to MnpExternalMapping. 'exact' may ONLY be set by a "
                        "human who read both concept definitions (brief §10, §11).",
                   widths={"external_id": 45, "external_label": 35, "note": 30})
    # column letters: I=review_state, H=proposed_mapping_type (0-indexed from A)
    add_dropdown(ws, "I", list(config.MNP_MAPPING_REVIEW_STATES), first_row=4)
    add_dropdown(ws, "H", list(config.MNP_MAPPING_TYPES) + ["unspecified"], first_row=4)


def _human_lab_templates(wb) -> None:
    add_table(wb, "55_PERSON_PROFILE_TEMPLATE",
              ["persona_id", "field", "value", "proficiency", "evidence", "note"],
              [["<slug>", "last_role", "", "", "", ""],
               ["<slug>", "skill: <name>", "", "strong", "verified", ""],
               ["<slug>", "constraint: <category>", "", "", "", "hardness: soft|hard"]],
              title="Person Profile — hand entry (mirror of the YAML in HUMAN_CLASSIFICATION_GUIDE.md)",
              note="The canonical authoring format is YAML under data_explorer/human_lab/examples/ or "
                   "data/data_explorer/human_lab/. This sheet is a scratch pad; run "
                   "`python -m data_explorer.cli golden` on the YAML files.")
    ws = wb["55_PERSON_PROFILE_TEMPLATE"]
    add_dropdown(ws, "D", sorted(hl_schema.PROFICIENCY), first_row=4)
    add_dropdown(ws, "E", sorted(hl_schema.EVIDENCE), first_row=4)

    add_table(wb, "58_EXPECTED_RESULT_TEMPLATE",
              ["persona_id", "career_code", "expected_feasibility", "expected_transition_distance", "verdict", "note"],
              [["<slug>", "sales_manager", "reachable", "d1_progression", "acceptable", ""]],
              title="Human Expected Result — hand entry",
              note="Canonical format is YAML (<persona>.expected.yaml).")
    ws = wb["58_EXPECTED_RESULT_TEMPLATE"]
    add_dropdown(ws, "C", sorted(hl_schema.FEASIBILITY), first_row=4)
    add_dropdown(ws, "D", sorted(hl_schema.TRANSITION), first_row=4)
    add_dropdown(ws, "E", ["expected_top", "acceptable", "unacceptable"], first_row=4)


def _golden_export(wb) -> None:
    import json
    rows = []
    for p in sorted(config.GOLDEN_OUT_DIR.glob("case_*.json")):
        case = json.loads(p.read_text(encoding="utf-8"))
        rows.append([case["case_id"], case["persona_id"], ", ".join(case["expected_top_careers"]),
                     ", ".join(case["acceptable_careers"]), ", ".join(case["unacceptable_careers"]),
                     len(case["expected_gaps"]), ", ".join(case["expected_blockers"]) or "(none)",
                     case["created_by"], case["created_at"]])
    add_table(wb, "60_GOLDEN_EXPORT",
              ["case_id", "persona_id", "expected_top", "acceptable", "unacceptable",
               "n_gaps", "blockers", "created_by", "created_at"],
              rows, title="Exported Human Expected Results (evals/golden_data_explorer/)",
              note="Full fixtures incl. the embedded person snapshot are the JSON files. "
                   "Versioned, never silently rewritten (MNP_GOLDEN_DATASET_V1).")
