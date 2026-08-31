"""Shared IO: downloads (sha256-verified, idempotent), tab/CSV/XLSX
readers, provenance helpers.

Download + tab-reader + sha256 logic lifted from commit 6d27bf3
(`scripts/onet_import/common.py` + `download_onet.py`) and generalised.
"""

from __future__ import annotations

import csv
import hashlib
import subprocess
import sys
import zipfile
from collections.abc import Iterator
from pathlib import Path

csv.field_size_limit(min(sys.maxsize, 2**31 - 1))  # ESCO altLabels fields are large


def log(msg: str) -> None:
    print(msg, file=sys.stderr, flush=True)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def download(url: str, dest: Path, *, use_curl: bool = False) -> str:
    """Idempotent: if `dest` exists and its recorded sha256 matches, skip.
    Returns the sha256. Writes `<dest>.sha256` alongside."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    sha_file = dest.with_suffix(dest.suffix + ".sha256")
    if dest.exists() and sha_file.exists():
        recorded = sha_file.read_text(encoding="utf-8").split()[0]
        if sha256(dest) == recorded:
            log(f"  cached  {dest.name}  ({recorded[:12]}...)")
            return recorded

    if not use_curl:
        try:
            import requests  # noqa: PLC0415 — optional, see requirements-datalab.txt

            log(f"  GET     {url}")
            with requests.get(url, stream=True, timeout=180) as r:
                r.raise_for_status()
                with open(dest, "wb") as fh:
                    for chunk in r.iter_content(1 << 20):
                        fh.write(chunk)
        except ImportError:
            use_curl = True
    if use_curl:
        log(f"  curl    {url}")
        subprocess.run(["curl", "-sSL", "--fail", "-o", str(dest), url], check=True)

    digest = sha256(dest)
    sha_file.write_text(f"{digest}  {dest.name}\n", encoding="utf-8")
    log(f"  saved   {dest.name}  sha256 {digest}")
    return digest


def extract_zip(zip_path: Path, target_dir: Path) -> None:
    target_dir.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(target_dir)


def read_tab(path: Path) -> Iterator[dict[str, str]]:
    """O*NET tab-delimited file -> dict rows keyed by the header line.
    UTF-8, tab-separated, single header, no quoting."""
    with path.open("r", encoding="utf-8", newline="") as fh:
        yield from csv.DictReader(fh, delimiter="\t")


def read_csv(path: Path) -> Iterator[dict[str, str]]:
    """ESCO CSV -> dict rows. Comma-separated, RFC-4180 quoting, multi-line
    quoted fields (altLabels)."""
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        yield from csv.DictReader(fh)


def read_xlsx(path: Path, sheet: str | None = None) -> Iterator[dict[str, str]]:
    """First sheet (or named sheet) of an .xlsx -> dict rows keyed by row 1.
    Needs openpyxl (requirements-datalab.txt)."""
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    for r in rows:
        yield {header[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(header)}
    wb.close()


def read_xlsx_rows(path: Path, sheet: str | None = None) -> Iterator[tuple]:
    """Raw cell tuples of a sheet, no header handling (caller decides)."""
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    for r in ws.iter_rows(values_only=True):
        yield r
    wb.close()


def xlsx_sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def split_esco_labels(raw: str) -> list[str]:
    """ESCO altLabels / hiddenLabels are newline-separated inside one CSV
    field."""
    return [s.strip() for s in (raw or "").splitlines() if s.strip()]
