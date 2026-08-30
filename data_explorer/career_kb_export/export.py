"""Build data/data_explorer/exports/MNP_CAREER_KB_V1.xlsx from the MNP
Career KB (via data_explorer.mnp_snapshot, read-only).
"""

from __future__ import annotations

import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Font

from data_explorer import config
from data_explorer.excel._sheet import add_dropdown, add_table
from data_explorer.human_lab import schema as hl
from data_explorer.io import log
from data_explorer.mnp_snapshot import load_mnp_careers

OUT = config.EXPORT_DIR / "MNP_CAREER_KB_V1.xlsx"
DATASET_VERSION = "v1"

# review_status: the alpha KB is 100% editorial; nothing has been through a
# separate curator review workflow yet -> "editorial" everywhere. The
# column exists so a real review lifecycle drops straight in.
_REVIEW_STATES = ["editorial", "reviewed", "needs_review", "rejected", "unknown"]
_SOURCE_TYPES = ["MNP_EDITORIAL", "OFFICIAL_UA", "ESCO", "ONET", "MARKET_SOURCE", "UNKNOWN"]


def _src_type(raw: str | None) -> str:
    if not raw:
        return "UNKNOWN"
    r = raw.lower()
    if "editorial" in r or r.startswith("mnp"):
        return "MNP_EDITORIAL"
    if "esco" in r:
        return "ESCO"
    if "onet" in r or "o*net" in r:
        return "ONET"
    if "official" in r or "classifier" in r or "dsz" in r:
        return "OFFICIAL_UA"
    return "UNKNOWN"


def _review(raw: str | None) -> str:
    return "editorial" if (raw and "editorial" in raw.lower()) else "unknown"


_DIFFICULTY_UK = {"easy": "Низька", "moderate": "Середня", "challenging": "Висока", "hard": "Дуже висока"}
_ENTRY_WO_EXP_UK = {
    "yes": "Так", "limited": "Частково", "no": "Ні", "unknown": "Немає підтверджених даних",
}
_IMPORTANCE_UK = {"low": "Низька", "medium": "Середня", "high": "Висока", "critical": "Критична"}
_REQ_TYPE_UK = {
    "must_have": "Обов'язкова", "high_value": "Дуже бажана",
    "differentiator": "Перевага", "optional": "Додатково",
}
_LEVEL_UK = {"basic": "Базовий", "working": "Впевнений", "strong": "Високий"}
_PROCON_UK = {"advantage": "Перевага", "disadvantage": "Недолік"}
_PATH_TYPE_UK = {
    "entry": "Старт", "junior": "Початковий", "core": "Основний",
    "senior": "Досвідчений", "lead": "Керівний", "executive": "Топ-рівень",
}
_HARDNESS_UK = {"soft": "Бажана", "hard": "Обов'язкова (підтверджено)"}


def build(dest=None, careers=None) -> None:
    from pathlib import Path
    dest = Path(dest) if dest else OUT
    if careers is None:
        careers = load_mnp_careers()
    wb = Workbook()
    wb.remove(wb.active)
    generated_at = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")

    _readme(wb, careers, generated_at)
    _careers(wb, careers)
    _skills(wb, careers)
    _requirements(wb, careers)
    _responsibilities(wb, careers)
    _career_paths(wb, careers)
    _pros_cons(wb, careers)
    _market_data(wb, careers)
    _external_refs(wb, careers)
    _provenance(wb, careers)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    log(f"  wrote {dest}  ({len(wb.sheetnames)} sheets, {len(careers)} careers)")


# --------------------------------------------------------------------------
def _readme(wb, careers, generated_at) -> None:
    ws = wb.create_sheet("00_README")
    lines = [
        ("MNP CAREER KNOWLEDGE BASE — V1 EXPORT", True),
        (f"generated_at: {generated_at}    dataset_version: {DATASET_VERSION}    careers: {len(careers)} ACTIVE", False),
        ("", False),
        ("The production MNP Career KB (mnp_* tables) is the SINGLE SOURCE OF TRUTH.", False),
        ("This workbook is a READ / REVIEW / ANALYSIS view. It is NOT a source of truth and there is NO Excel -> DB path.", False),
        ("Rebuild: python -m data_explorer.cli export-careers-excel", False),
        ("", False),
        ("Sheets", True),
        ("10_CAREERS         одна професія на рядок (mnp_careers + mnp_career_families) + складність/вхід", False),
        ("20_SKILLS          Career<->Skill (mnp_career_skill_requirements + mnp_skills), Тверді/М'які", False),
        ("30_REQUIREMENTS    освіта/досвід/мова/сертифікація/ліцензія/інші (mnp_career_requirements)", False),
        ("40_RESPONSIBILITIES  обов'язки (mnp_career_tasks — MNP_CAREER_PROFILE_SCHEMA_V1 §7)", False),
        ("50_CAREER_PATHS    типовий кар'єрний шлях (mnp_career_path_steps) — НЕ гарантований маршрут", False),
        ("60_PROS_CONS       переваги/недоліки (mnp_career_pros_cons) — РЕДАКЦІЙНИЙ шар MNP, не статистика", False),
        ("70_MARKET_DATA     (mnp_market_snapshots + mnp_salary_snapshots) — порожньо/UNKNOWN; ЖОДНИХ вигаданих цифр", False),
        ("80_EXTERNAL_REFS   ESCO/O*NET/ISCO/UA_CLASSIFIER (mnp_external_mappings, entity_type=career)", False),
        ("90_PROVENANCE      чому кожне значення в KB — source/source_version/confidence по кожному полю", False),
        ("", False),
        ("ПРАВИЛА: Ukrainian-first (українські колонки — перші). UNKNOWN != 0 (порожня клітинка). Без AI. "
         "Назви навичок — людські (ніколи не UUID).", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(bold=True, size=12, color="1F3864")
    ws.column_dimensions["A"].width = 120


def _careers(wb, careers) -> None:
    rows = [[
        c.canonical_name_uk, c.family_name_uk or c.family_code, c.description_short_uk,
        c.description_long_uk,
        _DIFFICULTY_UK.get(c.difficulty_level, "Немає підтверджених даних") if c.difficulty_level
        else "Немає підтверджених даних",
        _ENTRY_WO_EXP_UK.get(c.entry_without_experience, "Немає підтверджених даних"),
        c.typical_entry_route_uk,
        "Так" if c.market_data_limited else "Ні",
        c.code, c.canonical_name_en, c.status, c.career_profile_version, c.updated_at,
    ] for c in careers]
    add_table(wb, "10_CAREERS",
              ["Назва (укр)", "Категорія", "Короткий опис", "Повний опис",
               "Складність входу", "Старт без досвіду", "Типовий вхід (укр)", "Ринок обмежено",
               "career_code", "name_en", "status", "profile_version", "updated_at"],
              rows, title="Професії MNP (джерело: mnp_careers / mnp_career_families)",
              note="Ukrainian-first: усі кар'єрні картки читаються українською без англійських колонок. "
                   "«Складність входу» / «Старт без досвіду» — з mnp_careers. «Ринок обмежено»=Так -> ринкові "
                   "показники недоступні (жодних вигаданих цифр).",
              widths={"Короткий опис": 55, "Повний опис": 60, "Назва (укр)": 30, "Типовий вхід (укр)": 60,
                      "name_en": 28})


def _skills(wb, careers) -> None:
    SOFT = {"communication", "management"}
    rows = []
    for c in careers:
        for s in c.skill_requirements:
            group_uk = "М'яка" if (s["skill_type"] or "").lower() in SOFT else "Тверда"
            rows.append([
                c.canonical_name_uk, s["skill_uk"], group_uk,
                _REQ_TYPE_UK.get(s["requirement_type"], s["requirement_type"]),
                _LEVEL_UK.get(s["proficiency_level"], s["proficiency_level"]),
                _IMPORTANCE_UK.get(s["importance"], s["importance"]),
                c.code, s["skill_en"], s["skill_type"], s["requirement_type"], s["proficiency_level"],
                _src_type(s["source_type"]), _review(s["source_type"]),
            ])
    ws = add_table(wb, "20_SKILLS",
              ["Професія (укр)", "Навичка (укр)", "Тип навички", "Потрібність (укр)", "Рівень (укр)",
               "Важливість (укр)",
               "career_code", "skill_name_en", "skill_type", "requirement_type", "proficiency_level",
               "source_type", "review_status"],
              rows, title="Навички професій (джерело: mnp_career_skill_requirements + mnp_skills)",
              note="Ukrainian-first. «Тип навички» Тверда/М'яка виводиться з skill_type "
                   "(communication|management -> М'яка). Назви навичок — людські, не UUID. UNKNOWN != 0.",
              widths={"Навичка (укр)": 34, "Професія (укр)": 30, "skill_name_en": 30})
    add_dropdown(ws, "M", _REVIEW_STATES, first_row=4)


_REQ_CATEGORY_UK = {
    "education": "Освіта", "experience": "Досвід", "language": "Мова",
    "credential": "Сертифікація", "legal": "Ліцензія та дозволи", "other": "Інші вимоги",
}


def _requirements(wb, careers) -> None:
    rows = []
    for c in careers:
        for r in c.requirements:
            rows.append([
                c.canonical_name_uk,
                _REQ_CATEGORY_UK.get(r["category"], r["category"]),
                r["description"],
                _HARDNESS_UK.get(r["hardness"], r["hardness"]),
                "так" if r["hard_blocker"] else "ні",
                c.code, r["category"], r["requirement_name"], "yes" if r["hard_blocker"] else "",
                r["value"], _src_type(r["source_type"]), _review(r["source_type"]), r.get("country"),
            ])
    ws = add_table(wb, "30_REQUIREMENTS",
              ["Професія (укр)", "Категорія (укр)", "Вимога (укр)", "Обов'язковість (укр)",
               "Жорсткий блокер",
               "career_code", "category", "requirement_name", "hard_blocker", "value",
               "source_type", "review_status", "country"],
              rows, title="Вимоги професій (джерело: mnp_career_requirements)",
              note="Ukrainian-first. Категорія: Освіта|Досвід|Мова|Сертифікація|Ліцензія та дозволи|Інші "
                   "(RequirementCategory education|experience|language|credential|legal|other). «Жорсткий блокер»=так "
                   "лише коли hardness=HARD. UNKNOWN != «немає вимоги».",
              widths={"Вимога (укр)": 50, "Професія (укр)": 26})
    add_dropdown(ws, "L", _REVIEW_STATES, first_row=4)


def _responsibilities(wb, careers) -> None:
    rows = []
    for c in careers:
        for t in c.tasks:
            rows.append([
                c.canonical_name_uk, t["title_uk"] or t["title_en"], t.get("description") or "",
                _IMPORTANCE_UK.get(t["importance"], t["importance"]),
                c.code, t["responsibility_id"], t["importance"],
                _src_type(t["source"]), _review(t["source"]),
            ])
    ws = add_table(wb, "40_RESPONSIBILITIES",
              ["Професія (укр)", "responsibility", "Опис (укр)", "Важливість (укр)",
               "career_code", "responsibility_id", "importance", "source_type", "review_status"],
              rows, title="Обов'язки професій (джерело: mnp_career_tasks — MNP_CAREER_PROFILE_SCHEMA_V1 §7)",
              note="Ukrainian-first. «responsibility» — це title_uk (українська назва обов'язку).",
              widths={"responsibility": 44, "Опис (укр)": 55, "Професія (укр)": 26})
    add_dropdown(ws, "I", _REVIEW_STATES, first_row=4)


def _career_paths(wb, careers) -> None:
    rows = []
    for c in careers:
        for s in sorted(c.path_steps, key=lambda s: (s["path_code"], s["step_order"])):
            rows.append([
                c.canonical_name_uk, s["step_order"], s["step_name_uk"],
                _PATH_TYPE_UK.get(s["step_type"], s["step_type"]),
                s.get("typical_experience_text_uk") or "", s.get("description_uk") or "",
                "так" if s["is_current_career_step"] else "",
                c.code, s["path_code"], s["step_type"], _src_type(s["source"]), _review(s["source"]),
            ])
    add_table(wb, "50_CAREER_PATHS",
              ["Професія (укр)", "Крок №", "Назва кроку (укр)", "Рівень (укр)",
               "Типовий досвід (укр)", "Опис (укр)", "Поточний крок",
               "career_code", "path_code", "step_type", "source_type", "review_status"],
              rows, title="Типовий кар'єрний шлях (джерело: mnp_career_path_steps — Founder Decision §6)",
              note="Ukrainian-first. Впорядкований типовий маршрут, НЕ гарантований шлях просування. Крок "
                   "кар'єрного шляху ніколи не створює окрему професію MNP. «Поточний крок»=так позначає рівень, "
                   "що відповідає самій професії.",
              widths={"Назва кроку (укр)": 34, "Опис (укр)": 50, "Професія (укр)": 26})


def _pros_cons(wb, careers) -> None:
    rows = []
    for c in careers:
        for p in sorted(c.pros_cons, key=lambda p: (p["type"], p["sort_order"])):
            rows.append([
                c.canonical_name_uk, _PROCON_UK.get(p["type"], p["type"]), p["sort_order"],
                p["text_uk"], c.code, p["type"], _src_type(p["source"]), p["review_status"],
            ])
    ws = add_table(wb, "60_PROS_CONS",
              ["Професія (укр)", "Тип (укр)", "№", "Твердження (укр)",
               "career_code", "type", "source_type", "review_status"],
              rows, title="Переваги та недоліки (джерело: mnp_career_pros_cons — Founder Decision §5)",
              note="Ukrainian-first. РЕДАКЦІЙНИЙ шар MNP (source=mnp_editorial_v1), а не об'єктивна статистика. "
                   "type: advantage (Перевага) | disadvantage (Недолік).",
              widths={"Твердження (укр)": 70, "Професія (укр)": 26})
    add_dropdown(ws, "H", _REVIEW_STATES, first_row=4)


def _market_data(wb, careers) -> None:
    rows = []
    for c in careers:
        if not c.market_snapshots:
            rows.append([c.code, "UA", "", "", "", "", "", "", "", "", "", "MARKET_DATA_LIMITED",
                         "", "", ""])
            continue
        for m in c.market_snapshots:
            rows.append([
                c.code, m["country"], m["region"], "", m["vacancy_count"], m["salary_median"],
                m["salary_p25"], m["salary_p75"], m["currency"], m["demand_trend"], m["demand_trend"],
                m["data_quality"], m["remote_share"], m["source"], m["collected_at"],
            ])
    add_table(wb, "70_MARKET_DATA",
              ["career_code", "country", "region", "city", "vacancy_count", "salary_median",
               "salary_min", "salary_max", "currency", "demand_level", "trend", "data_quality",
               "remote_share", "source", "collected_at"],
              rows, title="Market data (source: mnp_market_snapshots + mnp_salary_snapshots)",
              note="NO fabricated numbers. The alpha KB has zero market snapshots -> every career is "
                   "MARKET_DATA_LIMITED and every metric is blank. `salary_min/max` map to the model's p25/p75; "
                   "`city` / `demand_level` have no model field yet.",
              widths={"source": 22})


def _external_refs(wb, careers) -> None:
    rows = []
    for c in careers:
        for em in c.external_mappings:
            rows.append([c.code, em["source_system"], em["external_id"], em["external_label"],
                         em["mapping_type"], "candidate", em["confidence"], "", ""])
    add_table(wb, "80_EXTERNAL_REFS",
              ["career_code", "external_system", "external_id", "external_label", "mapping_type",
               "mapping_status", "confidence", "reviewed_by", "reviewed_at"],
              rows, title="External references (source: mnp_external_mappings, entity_type=career)",
              note="external_system in {ESCO, ONET, ISCO, UA_CLASSIFIER}. mapping_type in {exact, close, broad, "
                   "narrow}. `mapping_status` / `reviewed_by` / `reviewed_at` have no model field yet — shown as "
                   f"'candidate' / blank. {'(no career external mappings in the current KB)' if not rows else ''}",
              widths={"external_id": 45, "external_label": 30})


def _provenance(wb, careers) -> None:
    """One row per (career, entity_type, entity_id, field) that carries a
    source — 'why is this value in the KB'."""
    rows = []

    def add(code, etype, eid, fields: dict, source, source_version, review):
        for fname, _val in fields.items():
            rows.append([code, etype, eid, fname, _src_type(source), source or "",
                         source_version or "", review])

    for c in careers:
        add(c.code, "career", c.id,
            {"short_description": 1, "long_description": 1, "category": 1,
             "difficulty_level": 1, "entry_without_experience": 1, "typical_entry_route_uk": 1},
            "mnp_editorial_v1", None, "editorial")
        for ps in c.path_steps:
            add(c.code, "career_path_step", ps["entity_id"],
                {"step_name_uk": 1, "description_uk": 1, "typical_experience_text_uk": 1},
                ps["source"], ps["source_version"], _review(ps["source"]))
        for pc in c.pros_cons:
            add(c.code, "career_procon", pc["entity_id"], {"text_uk": 1},
                pc["source"], pc["source_version"], _review(pc["source"]))
        for s in c.skill_requirements:
            add(c.code, "career_skill", s["entity_id"],
                {"importance": 1, "required_level": 1, "requirement_type": 1},
                s["source_type"], s["source_reference"], _review(s["source_type"]))
        for r in c.requirements:
            add(c.code, "career_requirement", r["entity_id"],
                {"description": 1, "hardness": 1, "value": 1},
                r["source_type"], r["source_version"], _review(r["source_type"]))
        for t in c.tasks:
            add(c.code, "career_task", t["entity_id"], {"title": 1, "importance": 1},
                t["source"], t["source_version"], _review(t["source"]))
        for at in c.attributes:
            add(c.code, "career_attribute", at["entity_id"], {f"{at['group']}.{at['key']}": 1},
                at["source"], None, _review(at["source"]))
        for em in c.external_mappings:
            add(c.code, "career_external_mapping", em["entity_id"],
                {"mapping_type": 1, "confidence": 1}, em["source_system"], em["source_version"], "candidate")
        for m in c.market_snapshots:
            add(c.code, "market_snapshot", m["entity_id"], {"vacancy_count": 1, "salary_median": 1},
                m["source"], m["source_version"], "market")

    rows.sort()
    add_table(wb, "90_PROVENANCE",
              ["career_code", "entity_type", "entity_id", "field_name", "source_type",
               "source_reference", "source_version", "review_status"],
              rows, title="Provenance — why each value is in the Career KB",
              note="Provenance is NOT a separate table in the model: every KB row carries its own "
                   "source / source_version / confidence. This sheet flattens them per field. "
                   "source_type in {MNP_EDITORIAL, OFFICIAL_UA, ESCO, ONET, MARKET_SOURCE, UNKNOWN}.",
              widths={"entity_id": 38, "source_reference": 20})
