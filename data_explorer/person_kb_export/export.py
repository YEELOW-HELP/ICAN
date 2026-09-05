"""Build data/data_explorer/exports/MNP_PERSON_KB_V1.xlsx from the canonical
Person KB. Reads the DB at `MNP_DATABASE_URL` (default: the local dev
SQLite). Ukrainian-first headers, English codes alongside for traceability.
"""

from __future__ import annotations

import asyncio
import datetime as dt
import os
from pathlib import Path

from data_explorer import config
from data_explorer.excel._sheet import add_table

OUT = config.EXPORT_DIR / "MNP_PERSON_KB_V1.xlsx"
SCHEMA_VERSION = "person_kb_base_v1"

_DEFAULT_DB = "sqlite+aiosqlite:///./data/dev/mnp_dev.sqlite"


async def _load_persons() -> list[dict]:
    url = os.environ.get("MNP_DATABASE_URL", _DEFAULT_DB)
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app.db.base import Base  # noqa: F401
    from app.db import (  # noqa: F401
        models, models_access, models_assessment, models_career_card, models_career_kb_mnp,
        models_crm, models_identity, models_knowledge, models_matching_mnp, models_person_kb,
        models_platform, models_profile,
    )
    from app.services.person_kb import service
    from app.services.person_kb.views import serialize_person

    engine = create_async_engine(url)
    sm = async_sessionmaker(engine, expire_on_commit=False)
    async with sm() as s:
        persons = await service.list_persons(s)
        out = [serialize_person(await service.get_person(s, p.id)) for p in persons]
    await engine.dispose()
    return out


def build(dest: Path | str | None = None, persons: list[dict] | None = None) -> Path:
    from openpyxl import Workbook

    dest = Path(dest) if dest else OUT
    if persons is None:
        persons = asyncio.run(_load_persons())

    wb = Workbook()
    wb.remove(wb.active)
    gen = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    _readme(wb, persons, gen)
    _persons(wb, persons)
    _rows(wb, persons, "20_EDUCATION", "educations",
          ["ПІБ", "Рівень", "Заклад", "Спеціальність / кваліфікація", "Рік початку", "Рік завершення",
           "Статус", "Стан підтвердження", "person_id", "row_id", "education_level", "source"],
          lambda p, e: [_name(p), e["education_level_uk"], e["institution_name"],
                        e["specialty_or_qualification"], e["start_year"], e["end_year"], e["status_uk"],
                        e["evidence_state_uk"], p["id"], e["id"], e["education_level"], e["source"]])
    _rows(wb, persons, "30_CREDENTIALS", "credentials",
          ["ПІБ", "Тип", "Назва", "Провайдер", "Видано", "Дійсний до", "Номер", "Стан підтвердження",
           "person_id", "row_id", "credential_type", "source"],
          lambda p, c: [_name(p), c["credential_type_uk"], c["title"], c["provider"], c["issue_date"],
                        c["expiry_date"], c["credential_number"], c["evidence_state_uk"], p["id"], c["id"],
                        c["credential_type"], c["source"]])
    _rows(wb, persons, "40_EXPERIENCE", "experiences",
          ["ПІБ", "Посада (сирий факт)", "Компанія", "Початок", "Завершення", "Зараз тут", "Обов'язки",
           "Досягнення", "Інструменти", "Стан підтвердження", "person_id", "row_id",
           "canonical_career_id", "source"],
          lambda p, x: [_name(p), x["raw_job_title"], x["company_name"], x["start_date"], x["end_date"],
                        x["is_current_uk"], x["responsibilities_description"], x["achievements"],
                        x["tools_used"], x["evidence_state_uk"], p["id"], x["id"],
                        x["canonical_career_id"], x["source"]])
    _rows(wb, persons, "50_ACTIVITIES", "activities",
          ["ПІБ", "Тип", "Назва", "Організація", "Роль", "Початок", "Завершення", "Опис", "Результат",
           "Стан підтвердження", "person_id", "row_id", "activity_type", "source"],
          lambda p, a: [_name(p), a["activity_type_uk"], a["title"], a["organization"], a["role"],
                        a["start_date"], a["end_date"], a["description"], a["result_or_achievement"],
                        a["evidence_state_uk"], p["id"], a["id"], a["activity_type"], a["source"]])
    _rows(wb, persons, "60_SKILLS_TOOLS", "skills",
          ["ПІБ", "Навичка / інструмент", "У таксономії?", "Рівень", "Років досвіду", "Востаннє (рік)",
           "Стан підтвердження", "person_id", "row_id", "canonical_skill_id", "custom_status", "source"],
          lambda p, s: [_name(p), s["raw_input"], s["custom_status_uk"], s["proficiency_uk"],
                        s["years_used"], s["last_used_year"], s["evidence_state_uk"], p["id"], s["id"],
                        s["canonical_skill_id"], s["custom_status"], s["source"]])
    _rows(wb, persons, "70_LANGUAGES", "languages",
          ["ПІБ", "Мова", "Рівень", "Сертифікат", "Стан підтвердження", "person_id", "row_id", "level", "source"],
          lambda p, l: [_name(p), l["language"], l["level_uk"], l["certificate"], l["evidence_state_uk"],
                        p["id"], l["id"], l["level"], l["source"]])
    _mobility(wb, persons)
    _rows(wb, persons, "85_DOCUMENTS", "documents",
          ["ПІБ", "Тип документа", "Файл", "Розмір", "Нотатка", "person_id", "row_id", "document_type"],
          lambda p, d: [_name(p), d["document_type_uk"], d["filename"], d["file_size"], d["note"],
                        p["id"], d["id"], d["document_type"]])
    _evidence(wb, persons)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def _name(p: dict) -> str:
    c = p["core"]
    return " ".join(x for x in (c["first_name"], c["last_name"]) if x)


def _readme(wb, persons, gen) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet("00_README")
    lines = [
        ("MNP PERSON KNOWLEDGE BASE — V1 EXPORT", True),
        (f"generated_at: {gen}    schema_version: {SCHEMA_VERSION}    persons: {len(persons)}", False),
        ("", False),
        ("Канонічна БД Person KB (mnp_persons + дочірні таблиці) — ЄДИНЕ джерело істини.", False),
        ("Цей файл — READ / REVIEW експорт. Шляху Excel -> БД НЕМАЄ. Правки — через Admin / API.", False),
        ("Перегенерувати: python -m data_explorer.cli export-persons-excel", False),
        ("", False),
        ("Person KB = ФАКТ-ФІРСТ кар'єрний профіль: що ми дійсно знаємо про людину.", False),
        ("НЕ психологічний портрет / RIASEC / Big Five / Work Values / тести здібностей.", False),
        ("", False),
        ("Стани підтвердження (evidence):", True),
        ("  Зі слів            — людина / адміністратор ввели вручну", False),
        ("  Підтверджено документом — прикріплено підтверджуючий документ", False),
        ("  Знайдено системою (не підтверджено) — парсер знайшов кандидата; це ще НЕ факт", False),
        ("  Підтверджено користувачем — людина явно підтвердила кандидата з резюме", False),
        ("", False),
        ("ПРАВИЛО: UNKNOWN != NO. «Немає даних» — це не «Ні». Кожен факт «так/ні» — три стани.", False),
        ("Сира назва посади (raw_job_title) — незмінний факт; маппінг на Career KB — окреме поле.", False),
    ]
    for i, (t, b) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=t)
        if b:
            c.font = Font(bold=True, size=12, color="1F3864")
    ws.column_dimensions["A"].width = 110


def _persons(wb, persons) -> None:
    rows = [[
        _name(p), p["core"]["phone"], p["core"]["email"], p["core"]["telegram_username"],
        p["core"]["city"], p["core"]["region"], p["core"]["country"], p["core"]["date_of_birth"],
        p["core"]["status_uk"], p["core"]["source_uk"],
        len(p["educations"]), len(p["experiences"]), len(p["activities"]), len(p["skills"]),
        len(p["languages"]),
        p["id"], p["core"]["status"], p["core"]["source"], p["core"]["profile_version"],
        p["created_at"], p["updated_at"],
    ] for p in persons]
    add_table(wb, "10_PERSONS",
              ["ПІБ", "Телефон", "Email", "Telegram", "Місто", "Область", "Країна", "Дата народження",
               "Статус", "Джерело", "Освіта", "Досвід", "Активності", "Навички", "Мови",
               "person_id", "status", "source", "profile_version", "created_at", "updated_at"],
              rows, title="Люди (джерело: mnp_persons)",
              note="Ukrainian-first. Один рядок = одна людина. DRAFT-профіль не активується автоматично.",
              widths={"ПІБ": 26, "Email": 26})


def _rows(wb, persons, sheet, coll, headers, row_fn) -> None:
    rows = [row_fn(p, r) for p in persons for r in p[coll]]
    add_table(wb, sheet, headers, rows, title=f"{sheet} (джерело: mnp_person_{coll})",
              note="Ukrainian-first. «Стан підтвердження» показує, наскільки надійний факт. "
                   "«Знайдено системою» — це кандидат з резюме, ще не підтверджений.",
              widths={"ПІБ": 24})


def _mobility(wb, persons) -> None:
    rows = [[
        _name(p), p["mobility"]["has_driver_license_uk"], p["mobility"]["driver_license_categories"],
        p["mobility"]["has_car_uk"], p["mobility"]["willing_to_relocate_uk"],
        ", ".join(p["mobility"]["work_geography_uk"]), p["mobility"]["work_format_uk"],
        p["id"], p["mobility"]["has_driver_license"], p["mobility"]["has_car"],
        p["mobility"]["willing_to_relocate"], p["mobility"]["work_format"],
    ] for p in persons]
    add_table(wb, "80_MOBILITY",
              ["ПІБ", "Посвідчення водія", "Категорії", "Автомобіль", "Готовність до переїзду",
               "Географія роботи", "Формат роботи", "person_id",
               "has_driver_license", "has_car", "willing_to_relocate", "work_format"],
              rows, title="Мобільність (джерело: mnp_persons)",
              note="Ukrainian-first. UNKNOWN != NO: «Немає даних» не означає «Ні».",
              widths={"ПІБ": 24})


def _evidence(wb, persons) -> None:
    rows = []
    for p in persons:
        for coll, etype in (("educations", "education"), ("credentials", "credential"),
                            ("experiences", "experience"), ("activities", "activity"),
                            ("skills", "skill"), ("languages", "language")):
            for r in p[coll]:
                rows.append([_name(p), etype, r.get("raw_input") or r.get("title") or r.get("raw_job_title")
                             or r.get("institution_name") or r.get("language") or "",
                             r["evidence_state_uk"], r["source"],
                             "так" if r.get("supporting_document_id") else "",
                             p["id"], r["id"], r["evidence_state"]])
    add_table(wb, "90_EVIDENCE",
              ["ПІБ", "Тип запису", "Значення", "Стан підтвердження", "Джерело", "Є документ",
               "person_id", "row_id", "evidence_state"],
              rows, title="Стан підтвердження по кожному факту",
              note="«Знайдено системою (не підтверджено)» != підтверджений факт. "
                   "Парсер / система ніколи не створюють довіреного факту автоматично.",
              widths={"ПІБ": 24, "Значення": 40})


if __name__ == "__main__":  # pragma: no cover
    print(build())
