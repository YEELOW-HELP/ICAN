"""PERSON KB BASE V1 -- models, raw-fact immutability, shared skill
taxonomy, evidence states, CV intake, seed safety, and HTTP security /
cross-user isolation.
"""

import io

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import func, select

from app.api.main import app
from app.core.security import create_access_token, hash_password
from app.db.models import AdminRole, AdminUser
from app.db.models_career_card import MnpSkill, SkillStatus, SkillType
from app.db.models_person_kb import (
    CustomSkillStatus,
    MnpPerson,
    MnpPersonExperience,
    MnpPersonSkillV1,
    PersonEvidenceState,
    PersonSource,
    PersonStatus,
    TriState,
)
from app.db.session import get_session
from app.services.person_kb import cv_intake, seed_demo, service

CV_TEXT = (tests_fixture := __import__("pathlib").Path(__file__).parent / "fixtures" / "persona_c_cv.txt")


# ----------------------------------------------------------------------
async def _person(session, **kw):
    return await service.create_person(session, first_name=kw.pop("first_name", "Тест"),
                                       source=PersonSource.ADMIN_MANUAL, **kw)


async def _active_skill(session, name_uk="Робота з CRM", name_en="CRM Software"):
    s = MnpSkill(canonical_name_en=name_en, canonical_name_uk=name_uk, skill_type=SkillType.TOOL,
                 status=SkillStatus.ACTIVE, taxonomy_version="t")
    session.add(s)
    await session.flush()
    return s


# --- §48 models -------------------------------------------------------
async def test_person_defaults_and_uuid(session):
    p = await _person(session, first_name="Олена", phone="+380 67 1", telegram_username="@olena")
    assert p.id and p.status == PersonStatus.DRAFT
    assert p.profile_version == 1
    assert p.telegram_username == "olena"  # normalized, no '@'
    assert p.has_car == TriState.UNKNOWN  # UNKNOWN, never a bare False


async def test_telegram_url_is_normalized(session):
    p = await _person(session, telegram_username="https://t.me/some_user")
    assert p.telegram_username == "some_user"


async def test_tristate_unknown_is_not_no(session):
    p = await _person(session)
    p = await service.update_person_core(session, p.id, has_driver_license="yes")
    assert p.has_driver_license == TriState.YES
    assert p.has_car == TriState.UNKNOWN  # still explicitly unknown -- not "no"
    assert p.willing_to_relocate == TriState.UNKNOWN


async def test_nested_collections_crud(session):
    p = await _person(session)
    p = await service.add_row(session, p.id, "educations",
                              {"education_level": "bachelor", "institution_name": "ЛНУ", "end_year": 2021})
    p = await service.add_row(session, p.id, "activities",
                              {"activity_type": "volunteering", "title": "Волонтер"})
    p = await service.add_row(session, p.id, "languages", {"language": "English", "level": "b2"})
    assert len(p.educations) == 1 and len(p.activities) == 1 and len(p.languages) == 1
    row_id = p.educations[0].id
    p = await service.update_row(session, p.id, "educations", row_id, {"institution_name": "КНУ"})
    assert p.educations[0].institution_name == "КНУ"
    p = await service.delete_row(session, p.id, "educations", row_id)
    assert len(p.educations) == 0


async def test_experience_requires_raw_job_title(session):
    p = await _person(session)
    with pytest.raises(service.PersonKbError):
        await service.add_row(session, p.id, "experiences", {"company_name": "X"})


# --- §49 raw facts ---------------------------------------------------
async def test_raw_job_title_is_immutable_by_blank(session):
    p = await _person(session)
    p = await service.add_row(session, p.id, "experiences",
                              {"raw_job_title": "Старший менеджер відділу продажів", "company_name": "A"})
    rid = p.experiences[0].id
    # an update that blanks the raw title must NOT wipe the fact
    p = await service.update_row(session, p.id, "experiences", rid,
                                 {"raw_job_title": "", "company_name": "B"})
    assert p.experiences[0].raw_job_title == "Старший менеджер відділу продажів"
    assert p.experiences[0].company_name == "B"


async def test_canonical_career_mapping_is_a_separate_field(session):
    p = await _person(session)
    p = await service.add_row(session, p.id, "experiences", {"raw_job_title": "Продажник"})
    row = p.experiences[0]
    assert row.raw_job_title == "Продажник"
    assert row.canonical_career_id is None  # mapping is separate, not an overwrite


# --- §50 skills (shared taxonomy) ----------------------------------
async def test_person_skill_uses_career_kb_canonical_skill(session):
    skill = await _active_skill(session)
    p = await _person(session)
    p = await service.add_skill(session, p.id, canonical_skill_id=str(skill.id))
    ps = p.skills[0]
    assert ps.canonical_skill_id == skill.id  # the SAME mnp_skills row Career KB uses
    assert ps.custom_status == CustomSkillStatus.CANONICAL
    assert ps.proficiency is None  # UNKNOWN != beginner -- not fabricated


async def test_custom_skill_stays_pending_review_never_silently_canonical(session):
    p = await _person(session)
    p = await service.add_skill(session, p.id, raw_input="Дуже рідкісний інструмент QZX")
    ps = p.skills[0]
    assert ps.canonical_skill_id is None
    assert ps.custom_status == CustomSkillStatus.PENDING_REVIEW
    # and it did NOT create a new canonical skill
    n = (await session.execute(select(func.count()).select_from(MnpSkill))).scalar()
    assert n == 0


async def test_no_parallel_person_skill_dictionary():
    """PersonSkill points at mnp_skills -- there is no `person_skills_dictionary`."""
    from app.db.base import Base
    tables = set(Base.metadata.tables)
    assert "mnp_person_skills_v1" in tables
    assert not any("dictionary" in t and "person" in t for t in tables)
    fk = {fk.column.table.name for fk in
          Base.metadata.tables["mnp_person_skills_v1"].c["canonical_skill_id"].foreign_keys}
    assert fk == {"mnp_skills"}


# --- §51 evidence --------------------------------------------------
async def test_evidence_states_default_self_reported(session):
    p = await _person(session)
    p = await service.add_row(session, p.id, "educations", {"institution_name": "X"})
    assert p.educations[0].evidence_state == PersonEvidenceState.SELF_REPORTED


async def test_system_detected_is_not_user_confirmed(session):
    p = await _person(session)
    content = CV_TEXT.read_bytes()
    cands = await cv_intake.extract_candidates(session, p.id, filename="cv.txt", content=content)
    # nothing persisted yet -- candidates are not facts
    assert len(p.experiences) == 0
    fresh = await service.get_person(session, p.id)
    assert len(fresh.skills) == 0 and len(fresh.educations) == 0
    # only the CV document is saved
    assert len(fresh.documents) == 1

    # confirm -> USER_CONFIRMED, source CV_CONFIRMED
    await cv_intake.apply_confirmed(session, p.id, {
        "educations": cands["educations"], "skills": cands["skills"][:1],
        "languages": cands["languages"]}, document_id=cands["document_id"])
    fresh = await service.get_person(session, p.id)
    assert fresh.skills and fresh.skills[0].evidence_state == PersonEvidenceState.USER_CONFIRMED
    assert fresh.skills[0].source == PersonSource.CV_CONFIRMED


async def test_admin_edit_source_is_recorded_not_document_supported(session):
    p = await _person(session)
    p = await service.add_row(session, p.id, "credentials", {"title": "Курс"})
    rid = p.credentials[0].id
    p = await service.update_row(session, p.id, "credentials", rid, {"provider": "Laba"},
                                 source=PersonSource.ADMIN_EDIT)
    assert p.credentials[0].source == PersonSource.ADMIN_EDIT
    assert p.credentials[0].evidence_state == PersonEvidenceState.SELF_REPORTED  # NOT document_supported


# --- §53 CV -------------------------------------------------------
async def test_cv_parse_failure_keeps_file_and_offers_manual(session):
    p = await _person(session)
    with pytest.raises(cv_intake.CvParseError) as exc:
        await cv_intake.extract_candidates(session, p.id, filename="empty.txt", content=b"   ")
    assert exc.value.document_id is not None  # file NOT lost
    fresh = await service.get_person(session, p.id)
    assert len(fresh.documents) == 1


# --- §55 seed safety --------------------------------------------
async def test_repeated_demo_seed_does_not_overwrite_manual_edit(session):
    r1 = await seed_demo.seed_demo_persons(session)
    assert r1["created"] == 2
    persons = await service.list_persons(session)
    pid = persons[0].id
    await service.update_person_core(session, pid, city="РУЧНА ПРАВКА")
    r2 = await seed_demo.seed_demo_persons(session)
    assert r2["created"] == 0
    fresh = await service.get_person(session, pid)
    assert fresh.city == "РУЧНА ПРАВКА"


# --- §9 activation gate ---------------------------------------------
async def test_partial_profile_does_not_auto_activate(session):
    p = await _person(session)
    assert service.activation_readiness(p)  # something missing
    with pytest.raises(service.PersonKbError):
        await service.activate_person(session, p.id)
    p = await service.add_row(session, p.id, "educations", {"institution_name": "ЛНУ"})
    p = await service.activate_person(session, p.id)
    assert p.status == PersonStatus.ACTIVE


# ======================================================================
# HTTP -- security / cross-user isolation (§52)
# ======================================================================
@pytest_asyncio.fixture
async def client(session_factory):
    async def override():
        async with session_factory() as s:
            yield s
    app.dependency_overrides[get_session] = override
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c
    app.dependency_overrides.clear()


async def _admin_headers(session_factory) -> dict:
    async with session_factory() as s:
        a = AdminUser(email="pk_admin@ican.dev", password_hash=hash_password("pw"),
                      role=AdminRole.ADMIN, full_name="PK Admin")
        s.add(a)
        await s.commit()
        await s.refresh(a)
        return {"Authorization": f"Bearer {create_access_token(a.id, a.role.value)}"}


async def test_anonymous_cannot_touch_person_kb(client):
    assert (await client.get("/v1/mnp/admin/persons")).status_code == 401
    assert (await client.get("/v1/mnp/me/person")).status_code == 422  # missing X-Mnp-User-Id header


async def test_user_a_cannot_access_user_b_profile(client):
    ua = (await client.post("/v1/mnp/session")).json()["user_id"]
    ub = (await client.post("/v1/mnp/session")).json()["user_id"]
    ha, hb = {"X-Mnp-User-Id": ua}, {"X-Mnp-User-Id": ub}

    await client.post("/v1/mnp/me/person", json={"first_name": "Аліса", "city": "Київ"}, headers=ha)
    await client.post("/v1/mnp/me/person/educations", json={"institution_name": "КНУ"}, headers=ha)

    # user B sees their OWN (empty) profile, never user A's
    rb = await client.get("/v1/mnp/me/person", headers=hb)
    assert rb.status_code == 200
    assert rb.json().get("person") is None or rb.json().get("core", {}).get("first_name") != "Аліса"

    # user B, given user A's row ids, still cannot edit them (routes are self-scoped)
    ra = await client.get("/v1/mnp/me/person", headers=ha)
    edu_id = ra.json()["educations"][0]["id"]
    rbad = await client.patch(f"/v1/mnp/me/person/educations/{edu_id}",
                              json={"institution_name": "ЗЛАМАНО"}, headers=hb)
    assert rbad.status_code in (403, 404)
    ra2 = await client.get("/v1/mnp/me/person", headers=ha)
    assert ra2.json()["educations"][0]["institution_name"] == "КНУ"


async def test_admin_can_list_create_edit_archive(client, session_factory):
    h = await _admin_headers(session_factory)
    created = (await client.post("/v1/mnp/admin/persons",
                                 json={"first_name": "Богдан", "phone": "+380 50 5"}, headers=h)).json()
    pid = created["id"]
    assert created["core"]["status"] == "draft"

    await client.post(f"/v1/mnp/admin/persons/{pid}/experiences",
                      json={"raw_job_title": "Технік"}, headers=h)
    got = (await client.get(f"/v1/mnp/admin/persons/{pid}", headers=h)).json()
    assert got["experiences"][0]["raw_job_title"] == "Технік"

    arch = (await client.post(f"/v1/mnp/admin/persons/{pid}/archive", headers=h)).json()
    assert arch["core"]["status"] == "archived"

    lst = (await client.get("/v1/mnp/admin/persons", headers=h)).json()
    assert any(r["id"] == pid for r in lst)


async def test_cv_flow_over_http(client):
    uid = (await client.post("/v1/mnp/session")).json()["user_id"]
    h = {"X-Mnp-User-Id": uid}
    files = {"file": ("cv.txt", CV_TEXT.read_bytes(), "text/plain")}
    up = await client.post("/v1/mnp/me/person/cv", files=files, headers=h)
    assert up.status_code == 200 and up.json()["parsed"] is True
    cand = up.json()["candidates"]
    # confirm only the languages
    conf = await client.post("/v1/mnp/me/person/cv/confirm",
                             json={"document_id": cand["document_id"],
                                   "confirmed": {"languages": cand["languages"]}}, headers=h)
    assert conf.status_code == 200
    body = conf.json()
    assert body["languages"] and body["languages"][0]["evidence_state"] == "user_confirmed"
    assert len(body["experiences"]) == 0  # not confirmed -> not persisted


# --- §43 Excel reflects the real DB --------------------------------
async def test_excel_reflects_db_after_edit(session, tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from app.services.person_kb.views import serialize_person
    from data_explorer.person_kb_export import export

    p = await _person(session, first_name="Ганна")
    p = await service.add_row(session, p.id, "experiences", {"raw_job_title": "Аналітик"})

    out1 = tmp_path / "a.xlsx"
    export.build(dest=out1, persons=[serialize_person(await service.get_person(session, p.id))])
    w1 = load_workbook(out1)
    rows1 = list(w1["40_EXPERIENCE"].iter_rows(values_only=True))
    assert any("Аналітик" in str(c) for r in rows1 for c in r)

    await service.update_row(session, p.id, "experiences", p.experiences[0].id,
                             {"company_name": "ТОВ Нова Компанія"})
    out2 = tmp_path / "b.xlsx"
    export.build(dest=out2, persons=[serialize_person(await service.get_person(session, p.id))])
    w2 = load_workbook(out2)
    rows2 = list(w2["40_EXPERIENCE"].iter_rows(values_only=True))
    assert any("ТОВ Нова Компанія" in str(c) for r in rows2 for c in r)
