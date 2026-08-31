"""MNP_PERSON_KB_V1.xlsx exporter -- offline, builds from a list of
serialized persons (no DB round-trip needed for the sheet assertions)."""

from __future__ import annotations

import pytest

pytest.importorskip("openpyxl")

from data_explorer.person_kb_export import export

REQUIRED_SHEETS = [
    "00_README", "10_PERSONS", "20_EDUCATION", "30_CREDENTIALS", "40_EXPERIENCE",
    "50_ACTIVITIES", "60_SKILLS_TOOLS", "70_LANGUAGES", "80_MOBILITY", "85_DOCUMENTS", "90_EVIDENCE",
]


def _person(name="Тест Тестовий"):
    fn, ln = name.split(" ", 1)
    return {
        "id": "11111111-1111-1111-1111-111111111111",
        "core": {"first_name": fn, "last_name": ln, "phone": "+380 67 1", "email": None,
                 "telegram_username": "t", "city": "Львів", "region": None, "country": None,
                 "date_of_birth": None, "status": "active", "status_uk": "Активний",
                 "source": "user_manual", "source_uk": "Заповнив користувач", "profile_version": 3,
                 "notes": None},
        "mobility": {"has_driver_license": "yes", "has_driver_license_uk": "Так",
                     "driver_license_categories": "B", "has_car": "unknown", "has_car_uk": "Немає даних",
                     "willing_to_relocate": "no", "willing_to_relocate_uk": "Ні",
                     "work_geography": ["remote"], "work_geography_uk": ["Віддалено"],
                     "work_format": "remote", "work_format_uk": "Віддалено"},
        "educations": [{"id": "e1", "education_level": "bachelor", "education_level_uk": "Бакалавр",
                        "institution_name": "ЛНУ", "specialty_or_qualification": "Маркетинг",
                        "start_year": 2018, "end_year": 2022, "status": "completed",
                        "status_uk": "Завершено", "evidence_state": "self_reported",
                        "evidence_state_uk": "Зі слів", "source": "user_manual",
                        "supporting_document_id": None}],
        "credentials": [],
        "experiences": [{"id": "x1", "raw_job_title": "Старший менеджер відділу продажів",
                         "company_name": "ТОВ Ромашка", "start_date": None, "end_date": None,
                         "is_current": "yes", "is_current_uk": "Так",
                         "responsibilities_description": "Ведення клієнтів", "achievements": None,
                         "tools_used": "CRM", "industry": None, "employment_type": None,
                         "canonical_career_id": None, "evidence_state": "user_confirmed",
                         "evidence_state_uk": "Підтверджено користувачем", "source": "cv_confirmed",
                         "supporting_document_id": "d1"}],
        "activities": [{"id": "a1", "activity_type": "volunteering", "activity_type_uk": "Волонтерство",
                        "title": "Волонтер", "organization": None, "role": None, "start_date": None,
                        "end_date": None, "description": None, "result_or_achievement": None,
                        "evidence_state": "self_reported", "evidence_state_uk": "Зі слів",
                        "source": "user_manual", "supporting_document_id": None}],
        "skills": [{"id": "s1", "canonical_skill_id": "c1", "raw_input": "Excel",
                    "custom_status": "canonical", "custom_status_uk": "У таксономії",
                    "proficiency": None, "proficiency_uk": "Немає даних", "years_used": None,
                    "last_used_year": None, "notes": None, "evidence_state": "self_reported",
                    "evidence_state_uk": "Зі слів", "source": "user_manual",
                    "supporting_document_id": None}],
        "languages": [{"id": "l1", "language": "English", "level": "b2", "level_uk": "B2",
                       "certificate": None, "evidence_state": "user_confirmed",
                       "evidence_state_uk": "Підтверджено користувачем", "source": "cv_confirmed",
                       "supporting_document_id": None}],
        "documents": [{"id": "d1", "document_type": "cv", "document_type_uk": "Резюме",
                       "filename": "cv.txt", "mime_type": "text/plain", "file_size": 100,
                       "note": None, "created_at": "2026-09-01T00:00:00"}],
        "created_at": "2026-09-01T00:00:00", "updated_at": "2026-09-01T00:00:00",
    }


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    from openpyxl import load_workbook
    out = tmp_path_factory.mktemp("pkb") / "MNP_PERSON_KB_V1.xlsx"
    export.build(dest=out, persons=[_person(), _person("Друга Особа")])
    return load_workbook(out)


def _rows(ws):
    if not ws.tables:
        return []
    import re
    ref = next(iter(ws.tables.values())).ref
    m = re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)", ref)
    top, bot = int(m.group(1)), int(m.group(2))
    data = list(ws.iter_rows(min_row=top, max_row=bot, values_only=True))
    hdr = [str(h) for h in data[0]]
    return [dict(zip(hdr, r)) for r in data[1:]]


def test_all_sheets_present_in_order(wb):
    assert wb.sheetnames == REQUIRED_SHEETS


def test_person_rows_match(wb):
    rows = _rows(wb["10_PERSONS"])
    assert len(rows) == 2
    assert all(r["Статус"] == "Активний" for r in rows)  # Ukrainian label


def test_headers_are_ukrainian(wb):
    for sn in ("10_PERSONS", "40_EXPERIENCE", "80_MOBILITY"):
        headers = [c.value for c in next(wb[sn].iter_rows(min_row=wb[sn].tables and 3 or 1))]
        joined = " ".join(str(h) for h in headers if h)
        import re
        assert re.search("[А-Яа-яЇїІіЄєҐґ]", joined), sn


def test_raw_job_title_verbatim_in_excel(wb):
    rows = _rows(wb["40_EXPERIENCE"])
    assert rows[0]["Посада (сирий факт)"] == "Старший менеджер відділу продажів"


def test_evidence_sheet_distinguishes_states(wb):
    rows = _rows(wb["90_EVIDENCE"])
    states = {r["Стан підтвердження"] for r in rows}
    assert "Зі слів" in states and "Підтверджено користувачем" in states


def test_unknown_shown_not_as_no(wb):
    rows = _rows(wb["80_MOBILITY"])
    assert rows[0]["Автомобіль"] == "Немає даних"  # UNKNOWN, not "Ні"


def test_no_formulas(wb):
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("=")


def test_readme_states_db_is_source_of_truth(wb):
    text = " ".join(str(c.value) for row in wb["00_README"].iter_rows() for c in row if c.value)
    assert "джерело істини" in text and "UNKNOWN != NO" in text
