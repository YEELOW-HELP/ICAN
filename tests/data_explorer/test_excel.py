"""BLOCK 3 — Excel workbook smoke tests. Needs the reference DB + openpyxl."""

from __future__ import annotations

import pytest

from data_explorer import config

pytest.importorskip("openpyxl")
pytestmark = pytest.mark.skipif(not config.REFERENCE_DB.exists(), reason="reference.sqlite not built")


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    from openpyxl import load_workbook
    from data_explorer.excel import workbook

    out = tmp_path_factory.mktemp("xlsx") / "MNP_ESCO_ONET_DATA_EXPLORER.xlsx"
    workbook.build(dest=out)
    return load_workbook(out)


def test_workbook_has_the_expected_sheets(wb):
    names = set(wb.sheetnames)
    for required in ["00_README", "01_SOURCES", "10_MNP_CAREERS", "11_ESCO_OCCUPATIONS",
                     "12_ONET_OCCUPATIONS", "13_CAREER_CROSSWALK", "20_ONET_RATINGS",
                     "22_ONET_WORK_STYLES_FULL", "40_CAREER_EXPLORER", "42_DATA_QUALITY",
                     "50_MAPPING_REVIEW", "60_GOLDEN_EXPORT"]:
        assert required in names, required


def test_data_sheets_are_excel_tables_with_frozen_header(wb):
    for sn in ("01_SOURCES", "11_ESCO_OCCUPATIONS", "20_ONET_RATINGS", "50_MAPPING_REVIEW"):
        ws = wb[sn]
        assert ws.tables, f"{sn}: no Excel Table"
        assert ws.freeze_panes is not None, f"{sn}: header not frozen"


def test_no_formulas_no_vba(wb):
    # openpyxl won't load a .xlsm as .xlsx; assert no cell holds a formula string
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("="), f"{sn}!{cell.coordinate} is a formula"


def test_mapping_review_sheet_has_review_dropdowns(wb):
    ws = wb["50_MAPPING_REVIEW"]
    dv_formulas = " ".join(dv.formula1 or "" for dv in ws.data_validations.dataValidation)
    assert "candidate" in dv_formulas and "confirmed" in dv_formulas
    assert "broad" in dv_formulas and "narrow" in dv_formulas


def test_onet_ratings_sheet_covers_all_families_for_the_focus_set(wb):
    ws = wb["20_ONET_RATINGS"]
    header = [c.value for c in next(ws.iter_rows())] if ws.max_row else []
    # find the table_key column and collect distinct values
    # (header is on the row after the title/note — scan for it)
    families = set()
    tk_col = None
    for row in ws.iter_rows(values_only=True):
        if tk_col is None:
            if row and "table_key" in row:
                tk_col = row.index("table_key")
            continue
        if row[tk_col]:
            families.add(row[tk_col])
    assert {"interest", "work_style", "knowledge", "ability", "work_activity", "work_context"} <= families


def test_mnp_careers_sheet_shows_unmapped(wb):
    ws = wb["10_MNP_CAREERS"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "UNMAPPED" in text  # alpha KB has no external mappings yet
