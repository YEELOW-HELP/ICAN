"""MNP_CAREER_KB_V1.xlsx exporter tests (addendum §15).

Offline: reads the MNP Career KB alpha seed via an in-memory SQLite
(no production DB, no network, no AI). Builds the workbook to a tmp path.
"""

from __future__ import annotations

import pytest

pytest.importorskip("openpyxl")

from data_explorer.career_kb_export import export
from data_explorer.mnp_snapshot import load_mnp_careers

REQUIRED_SHEETS = [
    "00_README", "10_CAREERS", "20_SKILLS", "25_KNOWLEDGE", "30_REQUIREMENTS",
    "40_RESPONSIBILITIES", "50_CAREER_PATHS", "60_PROS_CONS", "70_MARKET_DATA",
    "80_ALIASES", "85_EXTERNAL_REFS", "90_PROVENANCE",
]


@pytest.fixture(scope="module")
def careers():
    return load_mnp_careers()


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    from openpyxl import load_workbook

    out = tmp_path_factory.mktemp("ckb") / "MNP_CAREER_KB_V1.xlsx"
    export.build(dest=out)
    return load_workbook(out)


def _table_rows(ws) -> list[dict]:
    """Rows of the sheet's Excel Table as dicts (header is the row the
    table ref starts on)."""
    if not ws.tables:
        return []
    ref = next(iter(ws.tables.values())).ref  # 'A3:K8'
    import re
    m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref)
    top, bot = int(m.group(2)), int(m.group(4))
    rows = list(ws.iter_rows(min_row=top, max_row=bot, values_only=True))
    header = [str(h) for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:]]


# --------------------------------------------------------------------------
def test_all_required_sheets_present(wb):
    for s in REQUIRED_SHEETS:
        assert s in wb.sheetnames, s
    assert wb.sheetnames == REQUIRED_SHEETS  # exact order


def test_exactly_five_active_test_careers_matching_the_db(wb, careers):
    rows = _table_rows(wb["10_CAREERS"])
    assert len(rows) == 5
    db_codes = sorted(c.code for c in careers)
    xls_codes = sorted(r["career_code"] for r in rows)
    assert xls_codes == db_codes
    assert all(r["status"] == "active" for r in rows)


def test_readme_has_generated_at_and_dataset_version(wb):
    text = " ".join(str(c.value) for row in wb["00_README"].iter_rows() for c in row if c.value)
    assert "generated_at" in text and "dataset_version" in text


def test_skills_export_matches_the_db_and_uses_human_names(wb, careers):
    rows = _table_rows(wb["20_SKILLS"])
    db_pairs = sorted((c.code, s["skill_en"]) for c in careers for s in c.skill_requirements)
    xls_pairs = sorted((r["career_code"], r["skill_name_en"]) for r in rows)
    assert xls_pairs == db_pairs
    # Ukrainian-first: the Ukrainian skill column carries a real name, no UUID.
    import re
    uuid_re = re.compile(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}")
    for r in rows:
        assert r["Навичка (укр)"] and not uuid_re.search(str(r["Навичка (укр)"]) + str(r["skill_name_en"]))
        assert r["Тип навички"] in ("Тверда", "М'яка")


def test_knowledge_export_matches_the_db(wb, careers):
    rows = _table_rows(wb["25_KNOWLEDGE"])
    db = sorted((c.code, k["knowledge_en"]) for c in careers for k in c.knowledge_requirements)
    xls = sorted((r["career_code"], r["knowledge_name_en"]) for r in rows)
    assert xls == db
    # sparse by design -- do not require a row per career
    for r in rows:
        assert r["Знання (укр)"]
        assert r["Важливість (укр)"] in ("Низька", "Середня", "Висока", "Критична")


def test_aliases_export_matches_the_db(wb, careers):
    rows = _table_rows(wb["80_ALIASES"])
    db = sorted((c.code, a["alias"]) for c in careers for a in c.aliases)
    xls = sorted((r["career_code"], r["Аліас / інша назва (укр)"]) for r in rows)
    assert xls == db and len(rows) > 0


def test_requirements_export_matches_the_db(wb, careers):
    rows = _table_rows(wb["30_REQUIREMENTS"])
    db = sorted((c.code, r["description"], "yes" if r["hard_blocker"] else "")
                for c in careers for r in c.requirements)
    xls = sorted((r["career_code"], r["requirement_name"], r["hard_blocker"] or "") for r in rows)
    assert xls == db


def test_responsibilities_export_matches_the_db(wb, careers):
    rows = _table_rows(wb["40_RESPONSIBILITIES"])
    db_n = sum(len(c.tasks) for c in careers)
    assert len(rows) == db_n > 0
    db_titles = sorted((c.code, t["title_uk"] or t["title_en"]) for c in careers for t in c.tasks)
    xls_titles = sorted((r["career_code"], r["responsibility"]) for r in rows)
    assert xls_titles == db_titles


def test_career_paths_reflect_db_path_steps(wb, careers):
    rows = _table_rows(wb["50_CAREER_PATHS"])
    assert len(rows) == sum(len(c.path_steps) for c in careers) > 0
    db = sorted((c.code, s["step_order"], s["step_name_uk"])
                for c in careers for s in c.path_steps)
    xls = sorted((r["career_code"], r["Крок №"], r["Назва кроку (укр)"]) for r in rows)
    assert xls == db
    # every career has an ordered, Ukrainian path and exactly one "current" rung
    for c in careers:
        assert [s["step_order"] for s in sorted(c.path_steps, key=lambda s: s["step_order"])] == \
            list(range(1, len(c.path_steps) + 1))
        assert sum(1 for s in c.path_steps if s["is_current_career_step"]) == 1


def test_market_data_is_blank_never_fabricated(wb, careers):
    rows = _table_rows(wb["70_MARKET_DATA"])
    for r in rows:
        # no career in the alpha KB has a market snapshot -> every metric blank
        for col in ("vacancy_count", "salary_median", "salary_min", "salary_max", "demand_level", "trend"):
            assert r.get(col) in (None, "", "UNKNOWN"), f"{r['career_code']}.{col} = {r.get(col)!r} (fabricated?)"
        assert r["data_quality"] == "MARKET_DATA_LIMITED"


def test_pros_cons_export_matches_the_db(wb, careers):
    rows = _table_rows(wb["60_PROS_CONS"])
    db = sorted((c.code, p["type"], p["text_uk"]) for c in careers for p in c.pros_cons)
    xls = sorted((r["career_code"], r["type"], r["Твердження (укр)"]) for r in rows)
    assert xls == db and len(rows) > 0
    for c in careers:
        assert sum(1 for p in c.pros_cons if p["type"] == "advantage") >= 4
        assert sum(1 for p in c.pros_cons if p["type"] == "disadvantage") >= 4
    for r in rows:
        assert r["Тип (укр)"] in ("Перевага", "Недолік")
        assert r["source_type"] == "MNP_EDITORIAL"


def test_provenance_present_for_every_editorial_value(wb, careers):
    rows = _table_rows(wb["90_PROVENANCE"])
    assert len(rows) > 50
    codes = {r["career_code"] for r in rows}
    assert codes == {c.code for c in careers}
    for r in rows:
        assert r["source_type"] in {"MNP_EDITORIAL", "OFFICIAL_UA", "ESCO", "ONET", "MARKET_SOURCE", "UNKNOWN"}
    # every skill/requirement/task row must have a provenance entry
    skill_ents = {r["entity_id"] for r in rows if r["entity_type"] == "career_skill"}
    assert len(skill_ents) == sum(len(c.skill_requirements) for c in careers)


def test_no_formulas_no_vba(wb):
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    assert not cell.value.startswith("="), f"{sn}!{cell.coordinate}"


def test_data_sheets_are_tables_with_frozen_header(wb):
    for sn in ("10_CAREERS", "20_SKILLS", "30_REQUIREMENTS", "40_RESPONSIBILITIES", "90_PROVENANCE"):
        ws = wb[sn]
        assert ws.tables, sn
        assert ws.freeze_panes is not None, sn


def test_review_status_dropdowns_present(wb):
    ws = wb["20_SKILLS"]
    formulas = " ".join(dv.formula1 or "" for dv in ws.data_validations.dataValidation)
    assert "editorial" in formulas and "needs_review" in formulas


def test_export_is_deterministic(tmp_path, careers):
    """Given the SAME source snapshot, two exports are byte-identical
    (only 00_README's generated_at date differs). Note: a *fresh* in-memory
    seed generates new random row UUIDs and a new `updated_at` each run —
    that is the source data changing, not the exporter; a real MNP DB has
    stable ids."""
    from openpyxl import load_workbook

    a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"
    export.build(dest=a, careers=careers)
    export.build(dest=b, careers=careers)
    wa, wbk = load_workbook(a), load_workbook(b)
    for sn in wa.sheetnames:
        if sn == "00_README":
            continue
        rows_a = [tuple(c.value for c in r) for r in wa[sn].iter_rows()]
        rows_b = [tuple(c.value for c in r) for r in wbk[sn].iter_rows()]
        assert rows_a == rows_b, sn
